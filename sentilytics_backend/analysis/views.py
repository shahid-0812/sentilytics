from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from analysis.utils import sentiment_model,tfidf_vectorizer,pre,generateWordcloud
from django.core.exceptions import ObjectDoesNotExist #?
import traceback

from .models import BatchComment, Comment
from .serializers import BatchCommentSerializer,CommentSerializer,CorrectedSentimentSerializer
from sentilytics.settings import YOUTUBE_API_KEY

import pandas as pd
import re
from googleapiclient.discovery import build
from io import BytesIO
import base64
from openpyxl import Workbook
from datetime import datetime,timedelta
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

from django.utils.dateparse import parse_date
from django.db.models import Count, Q
from django.utils.timezone import make_aware


    
# -----------------------------------------------------------------------------------------------------------------------------------------
#single comment analysis
class SingleCommentAnalysis(APIView):
    def post(self,request):
        try:
            data = request.data
            if "text" not in data:
                return Response("Text to analyze is not provided",status=status.HTTP_400_BAD_REQUEST)
            original_text = data["text"]
            if not original_text.strip():
                return Response("Text field is empty",status=status.HTTP_400_BAD_REQUEST)
            cleaned_text = pre.clean_text(original_text)
            vec_text = tfidf_vectorizer.transform([cleaned_text])
            sentiment = sentiment_model.predict(vec_text)[0]
            score = sentiment_model.predict_proba(vec_text)[0]
            score = round(score[sentiment], 2)
            sentiment_map = {0:"negative",1:"neutral", 2:"positive"}
            sentiment=sentiment_map[sentiment]
            if not request.user.is_authenticated:
                return Response({"sentiment":sentiment,"score":score,"message": "Anonymous users cannot store data."}, status=status.HTTP_202_ACCEPTED)
            comment_data = {
                "user": request.user.id,
                "comment": original_text,
                "cleaned_text": cleaned_text,
                "sentiment":sentiment,
                "score": score
            }
            serializer = CommentSerializer(data=comment_data)

            if serializer.is_valid():
                serializer.save(user=request.user)
                return Response({"sentiment":serializer.data['sentiment'],"score":serializer.data['score']}, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response("Error generating analysis",status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request, pk):
        try:
            if not request.user.is_authenticated:
                return Response({"message": "Anonymous users cannot access."}, status=status.HTTP_403_FORBIDDEN)
            comment = Comment.objects.get(pk=pk,user=request.user)
        except Comment.DoesNotExist:
            return Response({"error": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)

        if comment.is_updated:
            return Response({"error": "Sentiment has already been corrected once"}, status=status.HTTP_400_BAD_REQUEST)

        corrected_sentiment = request.data.get("sentiment")
        
        # Update the original SingleComment object
        if corrected_sentiment not in ["positive", "negative", "neutral"]:
            return Response({"error": "Invalid sentiment value"}, status=status.HTTP_400_BAD_REQUEST)
        
        if comment.sentiment==corrected_sentiment:
            return Response({"error": "same sentiment value as predicted"}, status=status.HTTP_400_BAD_REQUEST)

        corrected_data = {
            "user": request.user.id,
            "comment":comment.id,
            "comment_text": comment.comment,
            "predicted_sentiment": comment.sentiment,
            "corrected_sentiment": corrected_sentiment
        }
        corrected_serializer = CorrectedSentimentSerializer(data=corrected_data)
        
        if corrected_serializer.is_valid():
            corrected_serializer.save()
            comment.is_updated =True
            comment.save()
            return Response(corrected_data, status=status.HTTP_200_OK)

        return Response(corrected_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
#get all single comments related to user
    def get(self,request):
        if not request.user.is_authenticated:
                return Response({"message": "Anonymous users cannot store data."}, status=status.HTTP_403_FORBIDDEN)
        user_comments = Comment.objects.filter(user=request.user,comment_type='single').order_by('-date_created')
        serializer = CommentSerializer(user_comments, many=True)
        return Response(serializer.data)

#delete single comment related to user
    def delete(self, request, pk):
        """Delete a specific comment by ID."""
        try:
            if not request.user.is_authenticated:
                return Response({"message": "Anonymous users cannot delete data."}, status=status.HTTP_403_FORBIDDEN)
            comment = Comment.objects.get(id=pk,user=request.user.id)
            comment.delete()
            return Response({"message": "Comment deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except Comment.DoesNotExist:
            return Response({"error": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# -----------------------------------------------------------------------------------------------------------------------------------------
#multiple comments analysis
class MultipleCommentsAnalysis(APIView):
    permission_classes=[IsAuthenticated]
    def is_date(self,value):
        try:
            parsed = pd.to_datetime(value, errors='coerce')  # Convert to datetime
            return not pd.isna(parsed)
        except:
            return False
    def post(self,request):
        try:
            # Ensure a file is uploaded
            if "file" not in request.FILES:
                return Response({"error": "CSV or Excel file is required."}, status=status.HTTP_400_BAD_REQUEST)

            file = request.FILES["file"]
            file_extension = file.name.split(".")[-1].lower()

            # Read file into DataFrame
            try:
                if file_extension == "csv":
                    df = pd.read_csv(file)
                    file_type = "CSV File"
                elif file_extension in ["xls", "xlsx"]:
                    df = pd.read_excel(file)
                    file_type = "Excel File"
                else:
                    return Response({"error": "Unsupported file format. Please upload a CSV or Excel file."}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": f"Invalid file format: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

            # Ensure the column is provided
            column = request.data.get("column")
            batchname = request.data.get("batchname")
            if not column:
                return Response({"error": "Column name is required."}, status=status.HTTP_400_BAD_REQUEST)
            
            if not batchname:
                return Response({"error": "Batch name is required."}, status=status.HTTP_400_BAD_REQUEST)
            
            if BatchComment.objects.filter(user=request.user.id,batchname=batchname).exists():
                return Response({"error": f"A batch with name {batchname} already exists."}, status=status.HTTP_400_BAD_REQUEST)
            # Ensure the column exists in the DataFrame
            if column not in df.columns:
                return Response({"error": f"File must contain a '{column}' column."}, status=status.HTTP_400_BAD_REQUEST)
            
            # Ensure the file is not empty
            if df.empty or df[column].dropna().empty:
                return Response({"error": "Uploaded file is empty or does not contain valid comments."}, status=status.HTTP_400_BAD_REQUEST)

            # Ensure the file has 5 or more comments
            if df[column].shape[0] < 5:
                return Response({"error": "Uploaded file contain less then 5 valid comments."}, status=status.HTTP_400_BAD_REQUEST)
            
            # Ensure the column has string values instead of numbers
            if df[column].apply(lambda x: isinstance(x, (int, float))).mean() > 0.8:
                return Response({"error": "The selected column appears to contain mostly numbers. Please provide a valid text column."}, status=status.HTTP_400_BAD_REQUEST)
            
            # Ensure the column has string values instead of dates
            if df[column].apply(self.is_date).mean()> 0.8:
                return Response({"error": "The selected column appears to contain mostly dates. Please provide a valid text column."}, status=status.HTTP_400_BAD_REQUEST)
            
            # Preprocess text
            df["cleaned_text"] = df[column].astype(str).apply(pre.clean_text)

            # Perform Sentiment Analysis
            vec_text = tfidf_vectorizer.transform(df["cleaned_text"])
            df["sentiment"] = sentiment_model.predict(vec_text)
            sentiment_map = {0: "negative",1:"neutral",2: "positive"}
            df[["score_neg","score_neu", "score_p"]] = sentiment_model.predict_proba(vec_text)
            df["score_neg"] = df["score_neg"].round(2)
            df["score_neu"] = df["score_neu"].round(2)
            df["score_p"] = df["score_p"].round(2)
            df["sentiment"] = df["sentiment"].map(sentiment_map)
            sentiment_counts = df["sentiment"].value_counts()
            
            # Sentiment Distribution Plot
            buff_word=generateWordcloud(df)
            Base64_word = base64.b64encode(buff_word.getvalue()).decode("utf-8")

            # Save Batch Analysis
            batch = BatchComment.objects.create(
                user=request.user,
                batchname=batchname,
                comment_type=file_type, 
                overall_sentiment=sentiment_counts.idxmax()
            )

            # Prepare comment objects for bulk creation
            comment_objects = [
                Comment(
                    user=request.user,
                    batch=batch,
                    comment=row[column],
                    cleaned_text=row["cleaned_text"],
                    sentiment=row["sentiment"],
                    score=row["score_p"] if row["sentiment"] == "positive" else row["score_neg"] if row["sentiment"]=="negative" else row["score_neu"],
                    comment_type='batch',
                )
                for _, row in df.iterrows()
            ]
            
            # Bulk save to database
            Comment.objects.bulk_create(comment_objects)
            sentiment_count=Comment.objects.filter(user=request.user.id,batch=batch).values("sentiment").annotate(count=Count("sentiment"))
            # Serialize data
            batch_serializer = BatchCommentSerializer(batch)
            comments_serializer = CommentSerializer(comment_objects, many=True)

            return Response(
                {
                    "batch_id": batch_serializer.data['id'],
                    "analyzed_comments": comments_serializer.data,
                    "BarChart": {s["sentiment"]: s["count"] for s in sentiment_count},
                    "wordcloud": "data:image/png;base64," + Base64_word,
                },
                status=201,
            )

        except KeyError as e:
            return Response({"error": f"Missing key: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError as e:
            return Response({"error": f"Invalid value: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        except ObjectDoesNotExist:
            return Response({"error": "Requested object does not exist."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            traceback.print_exc()  # Logs full traceback for debugging
            return Response({"error": f"An unexpected error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# -----------------------------------------------------------------------------------------------------------------------------------------
# Youtube comments analysis

class YoutubeCommentsAnalysis(APIView):
    permission_classes=[IsAuthenticated]
    def url_video_extract(self,url):
        pattern = r"(?:v=|\/)([0-9A-Za-z_-]{11}).*"
        match = re.search(pattern, url)
        return match.group(1) if match else None

    def post(self,request):
        video_url = request.data.get("vid_url")
        batchname = request.data.get("batchname")
        if not video_url:
            return Response({"error": "No YouTube URL provided"}, status=status.HTTP_400_BAD_REQUEST)

        video_id = self.url_video_extract(video_url)

        if not video_id:
            return Response({"error": "Invalid YouTube URL"}, status=status.HTTP_400_BAD_REQUEST)
        
        if not batchname:
            return Response({"error": "Batch name is required."}, status=status.HTTP_400_BAD_REQUEST)
            
        if BatchComment.objects.filter(user=request.user.id,batchname=batchname).exists():
            return Response({"error": f"A batch with name {batchname} already exists."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            youtube = build("youtube", "v3", developerKey=YOUTUBE_API_KEY)
            youtube_request = youtube.commentThreads().list(
                part="snippet",
                videoId=video_id,
                textFormat="plainText",
                maxResults=50,  # Fetch up to 50 comments
            )

            youtube_response = youtube_request.execute()
            comments = []

            for item in youtube_response.get("items", []):
                comment_text = item["snippet"]["topLevelComment"]["snippet"]["textDisplay"]
                comments.append(comment_text)

            # Convert list to DataFrame
            df = pd.DataFrame(comments, columns=["text"])
        except Exception as e:
            return Response({"error": f"Failed to fetch comments: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        df["cleaned_text"] = df["text"].astype(str).apply(pre.clean_text)
        vec_text = tfidf_vectorizer.transform(df["cleaned_text"])
        df["sentiment"] = sentiment_model.predict(vec_text)
        sentiment_map = {0: "negative",1:"neutral",2: "positive"}
        df[["score_neg","score_neu", "score_p"]] = sentiment_model.predict_proba(vec_text)
        df["score_neg"] = df["score_neg"].round(2)
        df["score_neu"] = df["score_neu"].round(2)
        df["score_p"] = df["score_p"].round(2)
        df["sentiment"] = df["sentiment"].map(sentiment_map)
        sentiment_counts = df["sentiment"].value_counts()
        
        #creating visuals
        buff_word=generateWordcloud(df)
        
        Base64_word = base64.b64encode(buff_word.getvalue()).decode("utf-8")
        batch = BatchComment.objects.create(user=request.user,comment_type="Youtube",overall_sentiment=sentiment_counts.idxmax(),batchname=batchname)
        
        comment_objects = [
            Comment(
                user=request.user,
                batch=batch,
                comment=row["text"],
                cleaned_text=row["cleaned_text"],
                sentiment=row["sentiment"],
                score=row["score_p"] if row["sentiment"] == "positive" else row["score_neg"] if row["sentiment"]=="negative" else row["score_neu"],
                comment_type='batch',
            )
            for _, row in df.iterrows()
        ]
        
        Comment.objects.bulk_create(comment_objects)
        sentiment_count=Comment.objects.filter(user=request.user.id,batch=batch).values("sentiment").annotate(count=Count("sentiment"))
        batch_serializer = BatchCommentSerializer(batch)
        comments_serializer = CommentSerializer(comment_objects, many=True)

        return Response(
            {
                "batch_id": batch_serializer.data['id'],
                "analyzed_comments": comments_serializer.data,
                "BarChart": {s["sentiment"]: s["count"] for s in sentiment_count},
                "wordcloud": "data:image/png;base64," + Base64_word,
            },
            status=201,
        )

# -----------------------------------------------------------------------------------------------------------------------------------------
#batch comments class
class Batch(APIView):

    permission_classes=[IsAuthenticated]
    def get(self,request, batch_id=None):
        # get all batches realated to the user
        if not batch_id:
            batches = BatchComment.objects.filter(user=request.user).order_by('-date_created')
            serializer = BatchCommentSerializer(batches, many=True)
            return Response(serializer.data)

        # In details batch comments
        try:
            batch = BatchComment.objects.get(id=batch_id, user=request.user)
        except BatchComment.DoesNotExist:
            return Response(
                {"error": "Batch not found or does not belong to the user."}, status=status.HTTP_404_NOT_FOUND
            )

        # Retrieve all related comments for the given batch
        comments = batch.comments.all()
        sentiment_count=Comment.objects.filter(user=request.user.id,batch_id=batch_id).values("sentiment").annotate(count=Count("sentiment"))
        # Serialize the comments data
        serializer = CommentSerializer(comments, many=True)
        df=pd.DataFrame(serializer.data)
        #creating wordcloud
        buff_word=generateWordcloud(df)
        Base64_word = base64.b64encode(buff_word.getvalue()).decode("utf-8")

        return Response(
            {
                "batch_id": batch.id,
                "batchname": batch.batchname,
                "comment_type": batch.comment_type,
                "date_created": batch.date_created,
                "comments": serializer.data,
                "BarChart": {s["sentiment"]: s["count"] for s in sentiment_count},
                "wordcloud": "data:image/png;base64," + Base64_word,
            }
        )

    def patch(self, request,batch_id, pk=None):
        try:
            batch=BatchComment.objects.get(id=batch_id,user=request.user)
            comment = batch.comments.get(pk=pk)
            
        except BatchComment.DoesNotExist:
            return Response({"error" : "Batch not found"},status=status.HTTP_404_NOT_FOUND)
        except Comment.DoesNotExist:
            return Response({"error": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)

        if comment.is_updated:
            return Response({"error": "Sentiment has already been corrected once"}, status=status.HTTP_400_BAD_REQUEST)

        corrected_sentiment = request.data.get("sentiment")
        
        # Update the original Singelomment object
        if corrected_sentiment not in ["positive", "negative", "neutral"]:
            return Response({"error": "Invalid sentiment value"}, status=status.HTTP_400_BAD_REQUEST)
        if comment.sentiment==corrected_sentiment:
            return Response({"error": "Same sentiment value as predicted"}, status=status.HTTP_400_BAD_REQUEST)
        corrected_data = {
            "user": request.user.id,
            "comment":comment.id,
            "comment_text": comment.comment,
            "predicted_sentiment": comment.sentiment,
            "corrected_sentiment": corrected_sentiment,
        }
        corrected_serializer = CorrectedSentimentSerializer(data=corrected_data)
        if corrected_serializer.is_valid():
            corrected_serializer.save()
            comment.is_updated = True
            comment.save()
            return Response(corrected_serializer.data, status=status.HTTP_200_OK)

        return Response(corrected_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    #delete batch comment related to user
    def delete(self, request, batch_id=None):
        """Delete a specific batch by ID."""
        try:
            Batch = BatchComment.objects.get(id=batch_id,user=request.user.id)
            Batch.delete()
            return Response({"message": "Batch deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except Batch.DoesNotExist:
            return Response({"error": "Batch not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DownloadAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, batch_id):
        # Fetch comments for the user and batch
        data = Comment.objects.filter(user=request.user, batch=batch_id)
        serialized_data = CommentSerializer(data, many=True)

        if not serialized_data.data:
            return Response({"error": "No data found"}, status=404)

        # Convert serialized data to DataFrame
        df = pd.DataFrame(serialized_data.data)

        date_created = df['date_created'].iloc[0] if not df.empty else "Unknown"
        date_created=datetime.fromisoformat(date_created).strftime('%B %d, %Y %I:%M %p')
        batchname=df['batchname'].iloc[0] if not df.empty else "Unknown"
        
        # Drop comment_type and date_created from the table
        df = df[['comment', 'sentiment', 'score']]

        # Create an in-memory Excel file using openpyxl
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sentiment Analysis"

        # Write title
        ws.append([f"Date: {date_created}"])
        ws.append([f"Batch Name:{batchname}"])
        ws.append([])

        # Write headers
        ws.append(["Comment", "Sentiment", "Score"])
        for cell in ws[ws.max_row]:  # Get the last row (header row)
            cell.font = Font(bold=True)
        # Write data rows
        for _, row in df.iterrows():
            ws.append(row.tolist())

        # Save the workbook
        wb.save(output)
        output.seek(0)

        # Create response with Excel file
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=sentiment_analayis_comments.xlsx'

        return response
    
    
class UserDashboardStats(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        try:
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("end_date")
            if start_date and end_date:
                parsed_start = parse_date(start_date)+timedelta(days=1)
                parsed_end = parse_date(end_date)+timedelta(days=1)

                if not parsed_start or not parsed_end:
                    return Response({"error": "Invalid date format"}, status=status.HTTP_400_BAD_REQUEST)

                start_date = make_aware(datetime.combine(parsed_start, datetime.min.time()))
                end_date = make_aware(datetime.combine(parsed_end, datetime.max.time()))
            else:
                start_date, end_date = None, None

            # Apply filters only if start_date and end_date are provided
            date_filter = Q()
            
            if start_date and end_date:
                date_filter &= Q(date_created__range=[start_date, end_date])
            

            # Fetch Data
            comment_count = Comment.objects.filter(date_filter,user=request.user.id).count() if Comment.objects.filter(date_filter,user=request.user.id).exists() else 0
            single_count = Comment.objects.filter(date_filter,user=request.user.id,comment_type='single').count() if Comment.objects.filter(date_filter,user=request.user.id,comment_type='single').exists() else 0
            batch_count = BatchComment.objects.filter(date_filter,user=request.user.id).count() if BatchComment.objects.filter(date_filter,user=request.user.id).exists() else 0
            sentiment_counts = Comment.objects.filter(date_filter,user=request.user.id).values("sentiment").annotate(count=Count("sentiment"))
            return Response(
                {
                    "total_comments": comment_count,
                    "total_single": single_count,
                    "total_batches": batch_count,
                    "sentiment_distribution": {s["sentiment"]: s["count"] for s in sentiment_counts},
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)